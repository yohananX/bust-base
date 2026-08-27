"""Import the Summer School 2026 XLSX into the Extra Lessons module.

The spreadsheet (Downloads/SUMMER SCHOOL 2026.xlsx) holds walk-in / external
summer pupils: pupil name, class, parent name, parent phone, amount paid,
gender. Per the Extra Lessons design, these children are imported as
*external* LessonEnrollment records (no Student/User portal accounts) — summer
pupils do not get a login.

The command is idempotent: re-running it will not create duplicate
enrollments, classes, or payments. It is safe to run against the dev DB and
against a production school (pass --school to target a specific tenant).

Sheet columns (Sheet1):
    NAME OF PUPIL/STUDENT | CLASS | PARENT'S NAME | PARENT'S PHONE NUMBER
    | AMOUNT PAID | GENDER

Notes / assumptions:
- A trailing "TOTAL" summary row is skipped.
- The per-pupil fee is uniform (₦5000); the sheet has no fee column, and the
  ₦90000 total equals 18 × ₦5000, confirming ₦5000 as the standard fee.
- Pupils with AMOUNT PAID = 0 are enrolled but left unpaid (UNPAID).
- Class names are normalised lightly (e.g. "Jss3" -> "Jss 3").
- Parent phone cells may contain two numbers; the first is used.
- Gender is not stored (LessonEnrollment has no gender field).
"""
from datetime import date
from decimal import Decimal, InvalidOperation

import re

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone

from core.models import School
from lessons.models import LessonEnrollment, LessonClass, LessonPeriod
from fees.models import Payment
from fees.paystack import issue_receipt

DEFAULT_PATH = r"C:\Users\pasto\Downloads\SUMMER SCHOOL 2026.xlsx"
DEFAULT_PERIOD = "Summer School 2026"
DEFAULT_FEE = "5000"


def _clean_text(value):
    return (value or "").strip()


def _clean_phone(value):
    """Strip quotes; take the first number when several are present."""
    if value is None:
        return ""
    raw = str(value).replace('"', "")
    parts = re.split(r"[,\n]", raw)
    return parts[0].strip() if parts else ""


def _normalize_class(value):
    s = _clean_text(value)
    if not s:
        return ""
    # Insert a space before a digit that follows a letter: "Jss3" -> "Jss 3".
    s = re.sub(r"([A-Za-z])(?=\d)", r"\1 ", s)
    return s.title()


def _to_decimal(value):
    if value is None or str(value).strip() == "":
        return Decimal("0.00")
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


class Command(BaseCommand):
    help = "Import Summer School 2026 XLSX into Extra Lessons (external enrollments)."

    def add_arguments(self, parser):
        parser.add_argument("--path", default=DEFAULT_PATH)
        parser.add_argument("--school", type=int, default=None,
                            help="School id to import into (defaults to first active school).")
        parser.add_argument("--period", default=DEFAULT_PERIOD)
        parser.add_argument("--fee", default=DEFAULT_FEE,
                            help="Uniform per-pupil fee (default 5000).")
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        path = options["path"]
        fee = _to_decimal(options["fee"])

        if options["school"]:
            school = School.objects.get(pk=options["school"])
        else:
            school = School.objects.filter(is_active=True).order_by("id").first()
        if not school:
            self.stderr.write("No school found to import into.")
            return

        period, created = LessonPeriod.objects.get_or_create(
            school=school, name=options["period"],
            defaults={
                "status": LessonPeriod.Status.OPEN,
                "start_date": date(2026, 8, 3),
                "end_date": date(2026, 8, 28),
            },
        )
        if created:
            self.stdout.write(f"Created period '{period.name}' (school={school.short_code}).")
        else:
            self.stdout.write(f"Reusing period '{period.name}' (school={school.short_code}).")

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        created_enrollments = 0
        existing_enrollments = 0
        created_payments = 0
        skipped_total = 0
        skipped_no_class = 0

        with transaction.atomic():
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = _clean_text(row[0])
                if not name:
                    continue
                if name.upper().startswith("TOTAL"):
                    skipped_total += 1
                    continue

                class_name = _normalize_class(row[1])
                if not class_name:
                    skipped_no_class += 1
                    continue

                parent_name = _clean_text(row[2]) or "—"
                phone = _clean_phone(row[3])
                parent_phones = [phone] if phone else []
                amount = _to_decimal(row[4])

                lesson_class, _ = LessonClass.objects.get_or_create(
                    school=school, period=period, name=class_name,
                    defaults={"fee_amount": fee},
                )

                enrollment, was_created = LessonEnrollment.objects.get_or_create(
                    school=school,
                    lesson_class=lesson_class,
                    external_name=name,
                    parent_phones=parent_phones,
                    defaults={
                        "parent_name": parent_name,
                        "status": LessonEnrollment.Status.REGISTERED,
                    },
                )
                if not was_created:
                    existing_enrollments += 1
                    continue

                created_enrollments += 1
                if amount > 0:
                    enrollment.status = (
                        LessonEnrollment.Status.PAID
                        if amount >= lesson_class.fee_amount
                        else LessonEnrollment.Status.REGISTERED
                    )
                    enrollment.save()
                    payment = Payment.objects.create(
                        school=school,
                        lesson_enrollment=enrollment,
                        student=None,
                        amount=amount,
                        method=Payment.Method.CASH,
                        status=Payment.Status.CONFIRMED,
                        paid_on=timezone.now(),
                        description=f"Summer School 2026 – {lesson_class.name}",
                    )
                    issue_receipt(payment)
                    created_payments += 1

            if options["dry_run"]:
                transaction.set_rollback(True)

        self.stdout.write(
            "\nImport summary:\n"
            f"  enrollments created : {created_enrollments}\n"
            f"  enrollments existing: {existing_enrollments}\n"
            f"  payments created    : {created_payments}\n"
            f"  skipped TOTAL row   : {skipped_total}\n"
            f"  skipped (no class)  : {skipped_no_class}\n"
            f"  fee per class       : {fee}"
        )
        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("DRY RUN — rolled back (no data written)."))
